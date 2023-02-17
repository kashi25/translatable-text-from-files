import os
import mimetypes
from openpyxl import Workbook, load_workbook
from docx import Document
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
import re
from builtins import PendingDeprecationWarning



# Scan the input folder for Excel, Word, and PDF files
input_folder = '/path/to/folder'
files = []
for filename in os.listdir(input_folder):
    file_path = os.path.join(input_folder, filename)
    if os.path.isfile(file_path) and mimetypes.guess_type(file_path)[0] in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/pdf'):
        files.append(file_path)

# Extract the translatable text from each file
output_wb = Workbook()
output_ws = output_wb.active
for file_path in files:
    try:
        file_type = mimetypes.guess_type(file_path)[0]
        if file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            wb = load_workbook(filename=file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            for sentence in cell.split('\n'):
                                formatted_sentence = ''
                                for char in sentence:
                                    # Check if the character is part of a formatting tag
                                    if char == '<':
                                        tag = ''
                                        while char != '>':
                                            tag += char
                                            char = next(sentence)
                                        tag += '>'
                                        # Add the tag to the formatted sentence
                                        formatted_sentence += tag
                                    # Check if the character is part of an inline icon
                                    elif char == '[':
                                        icon = ''
                                        while char != ']':
                                            icon += char
                                            char = next(sentence)
                                        icon += ']'
                                        # Add the icon to the formatted sentence
                                        formatted_sentence += '<icon>' + icon + '</icon>'
                                    else:
                                        # Add the character to the formatted sentence
                                        formatted_sentence += char
                                output_ws.append([file_path, formatted_sentence])
        elif file_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                formatted_text = ''
                for run in paragraph.runs:
                    # Add the run's text to the formatted text
                    formatted_text += run.text
                    # Check if the run has any formatting
                    if run.bold:
                        formatted_text += '<b/>'
                    if run.italic:
                        formatted_text += '<i/>'
                    if run.underline:
                        formatted_text += '<u/>'
                    if run.strike:
                        formatted_text += '<s/>'
                # Replace newlines with a placeholder character
                formatted_text = formatted_text.replace('\n', '|')
                # Find all inline icons and replace them with a placeholder tag
                formatted_text = re.sub(r'\[(.*?)\]', r'<icon>\1</icon>', formatted_text)
                # Find all other tags and replace them with a placeholder tag
                formatted_text = re.sub(r'<(.*?)>', r'<tag>\1</tag>', formatted_text)
                # Replace the placeholder character with a newline
                formatted_text = formatted_text.replace('|', '\n')
                # Replace the placeholder tags with the actual tags
                formatted_text = formatted_text.replace('<b>', '<b/>')
                formatted_text = formatted_text.replace('<i>', '<i/>')
                formatted_text = formatted_text.replace('<u>', '<u/>')
                formatted_text = formatted_text.replace('<s>', '<s/>')
                output_ws.append([file_path, formatted_text])
        elif file_type == 'application/pdf':
            text = extract_text(file_path)
        # Create a PDF canvas for drawing rectangles around text
        c = canvas.Canvas("temp_canvas.pdf")
        c.setStrokeColorRGB(1, 0, 0)
        for line in text.split('\n'):
            formatted_line = ''
            for char in line:
                # Check if the character is part of a formatting tag
                if char == '<':
                    tag = ''
                    while char != '>':
                        tag += char
                        char = next(line)
                    tag += '>'
                    # Add the tag to the formatted line
                    formatted_line += tag
                    # Check if the character is part of an inline icon
                elif char == '[':
                    icon = ''
                    while char != ']':
                        icon += char
                        char = next(line)
                    icon += ']'
                    # Add the icon to the formatted line
                    formatted_line += icon
                else:
                    # Add the character to the formatted line
                    formatted_line += char
                    # Draw a rectangle around each line of text
            c.rect(0, c._y, 1000, -10, fill=0)
            output_ws.append([file_path, formatted_line])
        c.save()
        # Delete the temporary canvas file
        os.remove("temp_canvas.pdf")
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")